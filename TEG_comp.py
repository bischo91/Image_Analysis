import sys
import os
from os import listdir
from os.path import isfile, join
import cv2
import openpyxl
from openpyxl.drawing.image import Image
import numpy as np
from matplotlib import pyplot as plt
from matplotlib import axes
import matplotlib.image as mpimg
from PIL import Image
import PIL
import math
import cv2
import skimage
from skimage.feature import hessian_matrix, hessian_matrix_eigvals
import skimage.feature
import skimage.viewer
from skimage.measure import regionprops
from scipy.ndimage import label
from skimage import measure
from skimage.morphology import skeletonize, medial_axis
import imutils
import math

def plot_panel(img_original, filename):
    arr = np.array(img_original)
    fig, ax = plt.subplots(1,2,figsize=(15,5), dpi=100)
    ax[0].imshow(img_original)
    ax[0].set_axis_off()
    ax[0].set_title(filename.replace('/',''))
    histR = cv2.calcHist([arr[:,:,0]], [0], None, [255], [0, 255])
    histG = cv2.calcHist([arr[:,:,1]], [0], None, [255], [0, 255])
    histB = cv2.calcHist([arr[:,:,2]], [0], None, [255], [0, 255])
    ax[1].plot(histR, 'r')
    ax[1].plot(histG, 'g')
    ax[1].plot(histB, 'b')
    ax[1].legend(['R', 'G', 'B'])
    ax[1].set_xlim([0, 255])
    ax[1].set_title('Rpeak: ' + str(histR[254]) + 'Gpeak: ' + str(histG[254]))
    ax[1].set_ylim([1, 250000])
    fig.tight_layout()


def feature_removal(img_in, eccentricity_min, area_max, area_min):
    labeled_array, num_features = label(img_in)
    properties = measure.regionprops(labeled_array)
    ecc = [prop.eccentricity for prop in properties]
    lb = [prop.label for prop in properties]
    area = [prop.area for prop in properties]
    bw = np.zeros(np.shape(img_in), dtype = bool)
    valid_label = set()
    ecce = []
    area = []
    for prop in properties:
        if (eccentricity_min < prop.eccentricity and (area_min < prop.area < area_max)):
            valid_label.add(prop.label)
            ecce.append(prop.eccentricity)
            area.append(prop.area)
    current_bw = np.in1d(labeled_array, list(valid_label)).reshape(np.shape(labeled_array))
    return current_bw

# Set current directory as path (where the py file is is the directory)
path = os.getcwd()
path = os.path.abspath("C:/Users/bisch/Downloads/after encap/TEG")


allfiles = [f for f in listdir(path) if isfile(join(path,f))]
imgfiles = []
for i in range(0, len(allfiles)):
    if allfiles[i].split('.')[1] in ['JPG']:
        imgfiles.append(allfiles[i])

for j in range(0, len(imgfiles)):
    filename = imgfiles[j]
    print(filename)
    # Load image
    img = mpimg.imread(path+'/'+filename)
    # Pixel Array
    arr = np.array(img)
    img = cv2.imread(path+'/'+filename)
    img = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)

    img_original = mpimg.imread(path+'/'+filename)

    sobely = cv2.Sobel(img, cv2.CV_64F, 1, 0, ksize=31)
    min = np.min(sobely)
    sobely = sobely - min
    max = np.max(sobely)
    div = max/255
    img = np.uint8(sobely/div)
    img = cv2.GaussianBlur(img, (9,9), 0)
    ret, img = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    # img = cv2.adaptiveThreshold(img, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 5, 0)
    # img = feature_removal(img, 0.97, 10000, 500)
    contours, h = cv2.findContours(img, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    x_list = []
    y_list = []
    coord = []
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area>2:
            approx = cv2.approxPolyDP(cnt, 0.05 * cv2.arcLength(cnt, True), True)
            # was 0.009
            if (len(approx) ==2):
                # cv2.drawContours(img_original, [approx], 0, (255,0,0), 2)
                corner_1 = approx[0]
                corner_2 = approx[1]
                x1 = corner_1[0][0]
                x2 = corner_2[0][0]
                y1 = corner_1[0][1]
                y2 = corner_2[0][1]
                x_list.append(x1)
                y_list.append(y1)
                x_list.append(x2)
                y_list.append(y2)

    x_min = np.min(x_list)
    y_max = np.max(y_list)
    y_min = np.min(y_list)
    x_max = np.max(x_list)
    img_copy = img_original.copy()

    x_a = np.max(x_list)
    y_a = y_list[x_list.index(x_a)]
    x_b = np.min(x_list)
    y_b = y_list[x_list.index(x_b)]
    y_c = np.max(y_list)
    x_c = x_list[y_list.index(y_c)]
    y_d = np.min(y_list)
    x_d = x_list[y_list.index(y_d)]

    h = np.sqrt((x_a-x_b)**2 + (y_a-y_b)**2)
    dy = abs(y_a-y_b)

    theta = np.arccos(dy/h)*180/np.pi
    if y_b-y_a < 0:
        theta = -theta

    img_copy = imutils.rotate(img_copy, theta)
    # for ii in range(x_min, x_max):
    #     for jj in range(y_min, y_max):
    #         if abs(ii-x_min) < 10 or abs(ii-x_max) <10:
    #             img_copy[jj,ii] = [255, 0, 0]
    #         if abs(jj-y_min) < 10 or abs(jj-y_max) <10:
    #             img_copy[jj,ii] = [255, 0, 0]
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = 'TEG #'
    sheet.cell(row=1, column=2).value = 'Average (5X5)'
    sheet.cell(row=1, column=3).value = 'Peak'
    fig, ax = plt.subplots(1,2)
    ax[0].set_axis_off()
    ax[0].set_title(filename)
    ax[1].set_axis_off()
    ax[0].imshow(img_copy)

    img = imutils.rotate(img_original, theta)
    img = img[y_min:y_max, x_min-50:x_max+50]
    img_cropped = img
    img_cropped_original = img_cropped

    img_cropped_GRAY = cv2.cvtColor(img_cropped, cv2.COLOR_RGB2GRAY)
    # img_cropped_GRAY = img_cropped_original

    img_cropped = cv2.cvtColor(img_cropped, cv2.COLOR_RGB2GRAY)

    line_max = []
    for k in range(0, np.shape(img_cropped)[0]):
        line_max.append(np.max(img_cropped[k]))
    max_avg = np.average(line_max)
    l = 0
    while np.max(img_cropped[l]) < 0.9*max_avg:
        img_cropped = np.delete(img_cropped, l, 0)
        l += 1
    k = np.shape(img_cropped)[0]-1
    while np.max(img_cropped[k]) < 0.9*max_avg:
        img_cropped = np.delete(img_cropped, k, 0)
        k -= 1

    new_y = np.shape(img_cropped)[0]
    new_x = np.shape(img_cropped)[1]
    step = round(np.shape(img_cropped)[0]/24)
    midpoint = round(step/2)


    ret, img_cropped = cv2.threshold(img_cropped, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)

    contours, h = cv2.findContours(img_cropped, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    x_list = []
    y_list = []
    coord = []
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area>50:
            approx = cv2.approxPolyDP(cnt, 0.1 * cv2.arcLength(cnt, True), True)
            if (len(approx) ==4):
                cv2.drawContours(img_cropped, [approx], 0, (255,0,0), 3)
    fig2, ax2 = plt.subplots(1,24)
    # ax2.imshow(img_cropped)

    TEG_position = []
    for i in range(0,24):
        TEG_position.append([])

    k = 0
    for i in range(1, new_y):
        # if np.max(img_cropped[i,:]) == 0:
            # switch = 'off'
        if np.max(img_cropped[i-1,:]) != np.max(img_cropped[i,:]):
            if np.max(img_cropped[i,:]) == 0:
                if k<23:
                    k+=1
        if np.max(img_cropped[i-1,:]) == np.max(img_cropped[i,:]):
            if np.max(img_cropped[i,:]) == 255:
                TEG_position[k].append(i)

    i=0
    for TEG in TEG_position:
        print(len(TEG))
        if TEG!=[]:
            y_min = np.min(TEG)
            y_max = np.max(TEG)
            img_TEG = img_cropped[y_min:y_max, 0:new_x]
            img_TEG_GRAY = img_cropped_GRAY[y_min:y_max, 0:new_x]

            k = 0
            l = np.shape(img_TEG)[1]-1
            while np.max(img_TEG[:, k]) == 0:
                k+=1
            while np.max(img_TEG[:,l]) == 0:
                l-=1

            x_min = k
            x_max = l
            img_TEG = img_TEG[:, k:l]
            img_TEG_GRAY = img_TEG_GRAY[:,k:l]
            # img_cropped_original[np.min(TEG),k:l] = [255,0,0]
            # img_cropped_original[np.max(TEG),k:l] = [255,0,0]
            # img_cropped_original[np.min(TEG):np.max(TEG),k] = [255,0,0]
            # img_cropped_original[np.min(TEG):np.max(TEG),l] = [255,0,0]
            # img_cropped_original[np.min(TEG):np.max(TEG),k:l] = [0,0,255]
            # ax2[i].imshow(img_cropped_original[np.min(TEG):np.max(TEG),k:l])
            ax2[i].imshow(img_cropped_GRAY[np.min(TEG):np.max(TEG),k:l])
            ax2[i].set_axis_off()
            i+=1

            pix_val_list = []

            for x in range(0, round((1/4)*np.shape(img_TEG)[1])):
                for y in range(0, np.shape(img_TEG)[0]):
                    if img_TEG[y, x] == 255:
                        pix_val_list.append(img_TEG_GRAY[y,x])
            for x in range(round((3/4)*np.shape(img_TEG)[1]),np.shape(img_TEG)[1]):
                for y in range(0, np.shape(img_TEG)[0]):
                    if img_TEG[y, x] == 255:
                        pix_val_list.append(img_TEG_GRAY[y,x])
            sheet.cell(row = i+1, column = 1).value = i
            sheet.cell(row = i+1, column = 2).value = np.average(pix_val_list)
            sheet.cell(row = i+1, column = 3).value = np.max(pix_val_list)
    plt.savefig(path+'/temp.png')
    # img_excel = cv2.imread(path+'/temp.png')
    img_temp = cv2.imread(path+'/temp.png')
    img_excel = openpyxl.drawing.image.Image(path+'/temp.png')
    # ratio = int(img_original.shape[0])/int(img_original.shape[1]) #H/W
    # W = 100
    img_excel.width = int(img_temp.shape[1])
    img_excel.height = int(img_temp.shape[0])

    # img_cropped_excel = openpyxl.drawing.image.Image(img_cropped)
    # ratio = int(img_original.shape[0])/int(img_original.shape[1]) #H/W
    # W = 100
    # img_cropped_excel.width = int(img_cropped.shape[1])*0.1
    # img_cropped_excel.height = int(img_cropped.shape[0])*0.1
    sheet.add_image(img_excel, 'E2')
    # sheet.add_image(img_cropped, 'E20')

    # print(TEG_position)
    # FIND OUT ALL  Y POSITIONS OF ON PIXLES
    #     print()xxx
                # print(approx)
    # for i in range(0, 24):q
    #     ax2[i,0].set_axis_off()
    #     ax2[i,1].set_axis_off()
    #     img_TEG = img_cropped[i*step:(i+1)*step, 0:new_x]
    #     img_cropped_original_TEG = img_cropped_original[i*step:(i+1)*step, 0:new_x]
    #     # img_TEG = cv2.cvtColor(img_TEG, cv2.COLOR_RGB2GRAY)
    #     peak = np.max(img_TEG[midpoint])
    #     peak_index = np.where(img_TEG==peak)
    #     y_index = peak_index[1]
    #     x_index = peak_index[0]
    #     y_center = y_index[round(len(y_index)/2)]
    #     x_center = x_index[round(len(x_index)/2)]
    #     y_extend = [j+2 for j in range(y_center-4, y_center+1)]
    #     x_extend_0 = [k+2 for k in range(x_center-4, x_center+1)]
    #     x_extend = []
    #     for x_pos in x_extend_0:
    #         if x_pos < np.shape(img_TEG)[0]:
    #             x_extend.append(x_pos)
    #     pixel_value_list=[]
    #     for x in x_extend:
    #         for y in y_extend:
    #             pixel_value_list.append(img_TEG[x,y])
    #             img_TEG[x, y] = 0
    #             img_cropped_original_TEG[x,y] = [255,0,0]
    #     sheet.cell(row = i+2, column = 1).value = i+1
    #     sheet.cell(row = i+2, column = 2).value = np.average(pixel_value_list)
    #     sheet.cell(row = i+2, column = 3).value = peak
    #     ax2[i, 0].imshow(img_cropped_original_TEG)
    #     ax2[i, 1].imshow(img_TEG)

    wb.save(path + '/' + filename.replace('jpg','xlsx').replace('JPG','xlsx'))
    ax[1].imshow(img_cropped)
plt.show()
os.remove(path+'/temp.png')
