import sys
import os
from os import listdir
from os.path import isfile, join
import cv2
import openpyxl
from openpyxl.drawing.image import Image
import numpy as np
from matplotlib import pyplot as plt
from matplotlib import axes
import matplotlib.image as mpimg
from PIL import Image
import PIL
import math
import cv2
import skimage
from skimage.feature import hessian_matrix, hessian_matrix_eigvals
import skimage.feature
import skimage.viewer
from skimage.measure import regionprops
from scipy.ndimage import label
from skimage import measure
from skimage.morphology import skeletonize, medial_axis
import imutils
import math
import tkinter.filedialog
import re


wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = 'TEG #'
sheet.cell(row=1, column=2).value = 'Average (5X5)'
sheet.cell(row=1, column=3).value = 'Peak'

def find_vg_from_filename(filename):
    plus = re.findall(r'%s(\d*\.?\d+)' % 'VG_\+', filename.upper())
    minus = re.findall(r'%s(\d*\.?\d+)' % 'VG_-', filename.upper())
    zero = re.findall(r'%s(\d*\.?\d+)' % 'VG_', filename.upper())
    if plus != []:
        return '+'+plus[0]
    elif minus != []:
        return '-'+minus[0]
    elif zero != []:
        return 0
    else:
        plus = re.findall(r'%s(\d*\.?\d+)' % 'VG \+', filename.upper())
        minus = re.findall(r'%s(\d*\.?\d+)' % 'VG -', filename.upper())
        zero = re.findall(r'%s(\d*\.?\d+)' % 'VG ', filename.upper())
        if plus != []:
            return '+'+plus[0]
        elif minus != []:
            return '-'+minus[0]
        elif zero != []:
            return 0
        else:
            return None

def rearrange_files_by_vg(fileinput):
    Vg_list = []
    for filename in fileinput:
        Vg_list.append(float(find_vg_from_filename(filename)))
    Vg_list_sorted = sorted(Vg_list)
    fileoutput = []
    for Vg in Vg_list_sorted:
        fileoutput.append(fileinput[Vg_list.index(Vg)])
    return fileoutput

def detect_TEG(img, img_original, dim):
    # Takes grey and original input and returns resized and rotated panel image and boolean whether panel is detected or not
    img = cv2.GaussianBlur(img, (25,25),0)
    ret, img = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    contours, h = cv2.findContours(img, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    x_max = 0
    y_max = 0
    x_min = math.inf
    y_min = math.inf
    counter = 0
    for count in contours:
        area = cv2.contourArea(count)
        if 5000>area>1000:
            counter += 1
            approx = cv2.approxPolyDP(count, 0.09 * cv2.arcLength(count, True), True)
            x1, y1 = approx[0][0]
            x2, y2 = approx[1][0]
            if max(x1, x2)>x_max:
                x_max = max(x1, x2)
                if x1>x2:
                    y_x_max = y1
                else:
                    y_x_max = y2
            if min(x1, x2)<x_min:
                x_min = min(x1, x2)
                if x1<x2:
                    y_x_min = y1
                else:
                    y_x_min = y2
            if max(y1, y2)>y_max:
                y_max = max(y1, y2)
                if y1>y2:
                    x_y_max = x1
                else:
                    x_y_max = x2
            if min(y1, y2)<y_min:
                y_min = min(y1, y2)
                if y1<y2:
                    x_y_min = x1
                else:
                    x_y_min = x2
    if 20<counter<26:
        x2 = x_max
        y2 = y_x_max
        x3 = x_y_max
        y3 = y_max
        x4 = x_min
        y4 = y_x_min
        x1 = x_y_min
        y1 = y_min
        hxy_1 = np.sqrt((x1-x2)**2 + (y1-y2)**2)
        hxy_2 = np.sqrt((x2-x3)**2 + (y2-y3)**2)
        hx = np.sqrt((x1-x2)**2 + (y1-y2)**2)
        if hxy_1>hxy_2:
            hx = hxy_2
            hy = hxy_1
            dh = hxy_1
            d = x1-x2
        else:
            hx = hxy_1
            hy = hxy_2
            dh = hxy_2
            d = y1-y2
        theta = np.arcsin(d/dh)*180/np.pi
        top_left_x = min([x1,x2,x3,x4])
        top_left_y = min([y1,y2,y3,y4])
        bot_right_x = max([x1,x2,x3,x4])
        bot_right_y = max([y1,y2,y3,y4])
    else:
        theta = dim[0]
        top_left_x = dim[1]
        top_left_y = dim[2]
        bot_right_x = dim[3]
        bot_right_y = dim[4]

    dy = math.ceil(abs(top_left_y-bot_right_y)*0.01)
    dx = math.ceil(abs(top_left_x-bot_right_x)*1)
    img_original = img_original[top_left_y-dy:bot_right_y+dy, top_left_x-dx:bot_right_x+dx]
    img_resized = imutils.rotate(img_original, theta)

    return img_resized, [theta, top_left_x, top_left_y, bot_right_x, bot_right_y]

def single_TEG(img_resized):

    line_max = []
    # Horizontal linear maximum
    for k in range(0, np.shape(img_resized)[0]):
        line_max.append(np.max(img_resized[k]))

    # average of maxima on each line
    # plt.figure()
    # plt.imshow(img_resized)
    fig_TEG, ax_TEG = plt.subplots(24, 1)

    max_avg = np.average(line_max)
    # Tighten vertical dimension
    l = 0
    while np.max(img_resized[l]) < 0.9*max_avg:
        img_resized = np.delete(img_resized, l, 0)
        l += 1
    k = np.shape(img_resized)[0]-1
    while np.max(img_resized[k]) < 0.9*max_avg:
        img_resized = np.delete(img_resized, k, 0)
        k -= 1
    # Define new dimension after tightening
    new_y = np.shape(img_resized)[0]
    new_x = np.shape(img_resized)[1]
    step = round(np.shape(img_resized)[0]/24)
    midpoint = round(step/2)

    # For each TEG, find max and get pixel values around it
    for i in range(0, 24):
        img_TEG = img_resized[i*step:(i+1)*step, 0:new_x]
        img_TEG_GRAY = cv2.cvtColor(img_TEG, cv2.COLOR_RGB2GRAY)
        # Define img_TEG_GRAY_temp after trimming img_TEG_GRAY by 4 to avoid finding maximum on edge
        img_TEG_GRAY_temp = img_TEG_GRAY[10:np.shape(img_TEG_GRAY)[0]-10, \
                            10:np.shape(img_TEG_GRAY)[1]-10]
        # Find maximum point
        peak = np.max(img_TEG_GRAY_temp[midpoint])
        # Find the index of the maximum
        peak_index = np.where(img_TEG_GRAY==peak)

        y_index = peak_index[1]
        x_index = peak_index[0]
        y_center = y_index[round(len(y_index)/2)]
        x_center = x_index[round(len(x_index)/2)]
        y_extend = [j+2 for j in range(y_center-4, y_center+1)]
        x_extend_0 = [k+2 for k in range(x_center-4, x_center+1)]
        x_extend = []
        for x_pos in x_extend_0:
            if x_pos < np.shape(img_TEG_GRAY)[0]:
                x_extend.append(x_pos)
        pixel_value_list=[]
        for x in x_extend:
            for y in y_extend:
                pixel_value_list.append(img_TEG_GRAY[x,y])
                # Mark plot with red
                img_TEG[x,y] = [255,0,0]

        img_TEG = np.asarray(img_TEG)
        # Show img_TEG with interested area marked
        ax_TEG[i].imshow(img_TEG)
        ax_TEG[i].set_axis_off()

        # Export to Excel
        sheet.cell(row = i+2, column = 1).value = i+1
        sheet.cell(row = i+2, column = 2).value = np.average(pixel_value_list)
        sheet.cell(row = i+2, column = 3).value = np.max(pixel_value_list)


# Set current directory as path (where the py file is is the directory)
path = os.getcwd()
path = os.path.abspath("C:/Users/bisch/Downloads/after encap/TEG")

root = tkinter.Tk()
path = tkinter.filedialog.askdirectory(parent=root, initialdir="/", title='Select Folder')
root.withdraw()

allfiles = [f for f in listdir(path) if isfile(join(path,f))]
allfiles = [f for f in allfiles if 'cropped' not in f and 'grid' not in f]
imgfiles = [f for f in allfiles if f.upper().endswith('.JPG')]
imgfiles = rearrange_files_by_vg(imgfiles)
prev_dim = [0, 0, 0, 0, 0]
fig, ax = plt.subplots(1, len(imgfiles))

for j in range(0, len(imgfiles)):
    filename = imgfiles[j]
    print(filename)
    # Load image
    img = mpimg.imread(path+'/'+filename)
    # Pixel Array
    arr = np.array(img)
    img = cv2.imread(path+'/'+filename)
    img_gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    img_original = mpimg.imread(path+'/'+filename)

    [img_resized, dim] = detect_TEG(img_gray, img_original, prev_dim)
    prev_dim = dim
    img_TEG = single_TEG(img_resized)

    new_img_file = path+'/' + filename.replace('.jpg','').replace('.JPG', '') + '_cropped.jpg'
    ax[j].imshow(img_resized)
    ax[j].set_axis_off()
    cv2.imwrite(new_img_file, cv2.cvtColor(img_resized, cv2.COLOR_BGR2RGB))
    wb.save(path + '/' + filename.replace('jpg','xlsx').replace('JPG','xlsx'))
plt.show()
