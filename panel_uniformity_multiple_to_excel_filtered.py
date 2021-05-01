import sys
import os
from os import listdir
from os.path import isfile, join
import cv2
import numpy as np
from matplotlib import pyplot as plt
from matplotlib import axes
import matplotlib.image as mpimg
from PIL import Image
import PIL
import math
import cv2
import re
import skimage
from skimage.feature import hessian_matrix, hessian_matrix_eigvals
import skimage.feature
import skimage.viewer
import imutils

import openpyxl
from openpyxl.drawing.image import Image


def plot_panel(img_original, filename):
    arr = np.array(img_original)
    fig, ax = plt.subplots(1,2,figsize=(15,5), dpi=100)
    ax[0].imshow(img_original)
    ax[0].set_axis_off()
    ax[0].set_title(filename.replace('/',''))

    histR = cv2.calcHist([arr[:,:,0]], [0], None, [255], [0, 255])
    histG = cv2.calcHist([arr[:,:,1]], [0], None, [255], [0, 255])
    histB = cv2.calcHist([arr[:,:,2]], [0], None, [255], [0, 255])

    # print(max(histR))
    ax[1].plot(histR, 'r')
    ax[1].plot(histG, 'g')
    ax[1].plot(histB, 'b')
    ax[1].legend(['R', 'G', 'B'])
    ax[1].set_xlim([0, 255])
    ax[1].set_title('Rpeak: ' + str(histR[254]) + 'Gpeak: ' + str(histG[254]))
    ax[1].set_ylim([1, 250000])
    fig.tight_layout()


def detect_panel(img, img_original):
    # Takes grey and original input and returns resized and rotated panel image and boolean whether panel is detected or not
    img = cv2.GaussianBlur(img, (9,9),0)
    ret, img = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    contours, h = cv2.findContours(img, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area>1000000:
            approx = cv2.approxPolyDP(cnt, 0.09 * cv2.arcLength(cnt, True), True)
    # was 0.009
            if (len(approx) == 4):
                # cv2.drawContours(img_original, [approx], 0, (255,0,0), 25)
                corners = approx

    try: corners
    except NameError: corners = None
    if corners is not None:
        x1, y1 = corners[0][0]
        x2, y2 = corners[1][0]
        x3, y3 = corners[2][0]
        x4, y4 = corners[3][0]

        hxy_1 = np.sqrt((x1-x2)**2 + (y1-y2)**2)
        hxy_2 = np.sqrt((x2-x3)**2 + (y2-y3)**2)
        hx = np.sqrt((x1-x2)**2 + (y1-y2)**2)

        if hxy_1>hxy_2:
            hx = hxy_2
            hy = hxy_1
            dh = hxy_1
            d = x1-x2
        else:
            hx = hxy_1
            hy = hxy_2
            dh = hxy_2
            d = y1-y2
        theta = np.arcsin(d/dh)*180/np.pi
        top_left_x = min([x1,x2,x3,x4])
        top_left_y = min([y1,y2,y3,y4])
        bot_right_x = max([x1,x2,x3,x4])
        bot_right_y = max([y1,y2,y3,y4])
        img_original = img_original[top_left_y:bot_right_y, top_left_x:bot_right_x]
        nx = np.shape(np.array(img_original))[1]
        ny = np.shape(np.array(img_original))[0]
        resize_dim_x_1 = math.ceil((nx - hx)/2)
        resize_dim_x_2 = resize_dim_x_1 + math.floor(hx)
        resize_dim_y_1 = math.ceil((ny - hy)/2)
        resize_dim_y_2 = resize_dim_y_1 + math.floor(hy)
        img_original = imutils.rotate(img_original, theta)
        img_resized = img_original[resize_dim_y_1:resize_dim_y_2, resize_dim_x_1:resize_dim_x_2]
        img_detect = True
    else:
        img_resized = []
        img_detect = False
    return img_resized, img_detect

def find_vg_from_filename(filename):
    plus = re.findall(r'%s(\d+)' % 'VG_\+', filename.upper())
    minus = re.findall(r'%s(\d+)' % 'VG_-', filename.upper())
    zero = re.findall(r'%s(\d+)' % 'VG_', filename.upper())
    if plus != []:
        return '+'+plus[0]
    elif minus != []:
        return '-'+minus[0]
    elif zero != []:
        return 0
    else:
        return None


# Set current directory as path (where the py file is is the directory)

# Set path
# path = os.path.abspath("C:/CS/python_ruby/image_process/RGB/Panel/Uniformity/QVGA 17/exposure time 1_60")
# path = os.path.abspath("C:/Users/bisch/Documents/Mattrix/QVGA Panel/JSR QVGA Panel/JSR QVGA #8_sprayed/photos/after encap_V_en_4V, V_Scan_4V, V_Data_Off_4V/Red")
# pathstr = r"C:\CS\python_ruby\image_process\RGB\Panel\Test Images\Green".replace("\\","/")
# pathstr = r"C:\Users\bisch\Documents\Mattrix\QVGA Panel\JSR QVGA Panel\JSR QVGA #12_sprayed\after encap\B".replace("\\","/")
pathstr = r"C:\Users\bisch\Documents\Mattrix\QVGA Panel\JSR QVGA Panel\JSR QVGA #11_sprayed\Photos after encap\B".replace("\\","/")


path = os.path.abspath(pathstr)
# Read all files in the folder
allfiles = [f for f in listdir(path) if isfile(join(path,f))]
allfiles = [f for f in allfiles if 'cropped' not in f and 'grid' not in f]
imgfiles = [f for f in allfiles if f.upper().endswith('.JPG')]


wb = openpyxl.Workbook()
sheet_overall = wb.active
k=0
for j in range(0, len(imgfiles)):
    filename = imgfiles[j]
    # Load image
    img = mpimg.imread(path+'/'+filename)
    # Pixel Array
    arr = np.array(img)
    img = cv2.imread(path+'/'+filename)
    img = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    img_original = mpimg.imread(path+'/'+filename)
    [img_resized, img_detect] = detect_panel(img, img_original)



    if img_detect == False or 'grid' in filename or 'cropped' in filename:
        print(filename + ': Detection Fail')
    else:
        print(filename + ': Complete')
        with_no_grid = path+'/' + filename.replace('.jpg','').replace('.JPG', '') + '_cropped.jpg'
        cv2.imwrite(with_no_grid, cv2.cvtColor(img_resized, cv2.COLOR_RGB2BGR))
        img_gray = cv2.cvtColor(img_resized, cv2.COLOR_RGB2GRAY)
        arr = np.array(img_gray)
        l = np.shape(np.array(arr))[0]
        w = np.shape(np.array(arr))[1]
        # print(str(w)+ 'X' + str(l) + '\n' + str(l/w))
        grid_x = 7
        grid_y = 9
        n = grid_y+1
        m = grid_x+1
        grid_size = 50
        dl = l/n
        dw = w/m
        coordinate_list =[]

        pixel_average_list = []
        pixel_std_list = []
        pixel_data = []
        # for grid

        for p in range(1, n):
            for q in range(1, m):
                # position_x.append(q*dw)
                coordinate_list.append([p*dl, q*dw])
                x_center = round(p*dl)
                y_center = round(q*dw)
                x_max = int(x_center + (grid_size/2))
                x_min = int(x_center - (grid_size/2))
                y_max = int(y_center + (grid_size/2))
                y_min = int(y_center - (grid_size/2))
                pixel_value_list = []
                # for region
                for ii in range(x_min, x_max):
                    for jj in range(y_min, y_max):
                        pixel_value_list.append(img_gray[ii,jj])
                        if abs(ii-x_min) < 5 or abs(ii-x_max) <5:
                            img_resized[ii,jj] = [255, 0, 0]
                        if abs(jj-y_min) < 5 or abs(jj-y_max) <5:
                            img_resized[ii,jj] = [255, 0, 0]
                grid_num = str(p) + 'X' + str(q)
                pixel_data.append([grid_num, np.average(pixel_value_list), np.std(pixel_value_list)])

        new_img_file = path+'/' + filename.replace('.jpg','').replace('.JPG', '') + '_grid.jpg'
        cv2.imwrite(new_img_file, cv2.cvtColor(img_resized, cv2.COLOR_RGB2BGR))


        Vg = find_vg_from_filename(filename)
        if Vg == None:
            Vg = filename
        sheet = wb.create_sheet('Vg = ' + str(Vg) + 'V')

        sheet.cell(row=1, column=1).value = 'Pixel #'
        sheet.cell(row=1, column=2).value = 'Average'
        sheet.cell(row=1, column=3).value = 'STDEV'

        for i in range(len(pixel_data)):
            sheet.cell(row=i+2, column=1).value = pixel_data[i][0]
            sheet.cell(row=i+2, column=2).value = pixel_data[i][1]
            sheet.cell(row=i+2, column=3).value = pixel_data[i][2]
        sheet.cell(row=1, column=5).value = 'Overall'
        sheet.cell(row=2, column=5).value = 'Average'
        sheet.cell(row=3, column=5).value = 'STDEV'
        sheet.cell(row=4, column=5).value = 'Lmin'
        sheet.cell(row=5, column=5).value = 'Lmax'
        sheet.cell(row=6, column=5).value = '100% - STDEV'
        sheet.cell(row=7, column=5).value = '1-(STDEV/AVG)'
        sheet.cell(row=9, column=5).value = 'Filtered'
        sheet.cell(row=10, column=5).value = 'Average'
        sheet.cell(row=11, column=5).value = 'STDEV'
        sheet.cell(row=12, column=5).value = 'Lmin'
        sheet.cell(row=13, column=5).value = 'Lmax'
        sheet.cell(row=14, column=5).value = '100% - STDEV'
        sheet.cell(row=15, column=5).value = '1-(STDEV/AVG)'


        pixel_average = np.average([j[1] for j in pixel_data])
        pixel_std = np.std([j[1] for j in pixel_data])
        pixel_median = np.median([j[1] for j in pixel_data])
        pixel_average_filtered = []
        pixel_number_filtered = []
        pixel_std_filtered = []
        not_counting = []

        pct_rng = 100
        for j in pixel_data:
            if pixel_median*((100-pct_rng)/100) < j[1] < pixel_median*((100+pct_rng)/100) \
                and j[2]<2*np.average([i[2] for i in pixel_data]) and j[1]*1.5>j[2]:
                pixel_number_filtered.append(j[0])
                pixel_average_filtered.append(j[1])
                pixel_std_filtered.append(j[2])
            else:
                not_counting.append(j[0])

        sheet.cell(row=2, column=6).value = pixel_average
        sheet.cell(row=3, column=6).value = pixel_std
        sheet.cell(row=4, column=6).value = min([j[1] for j in pixel_data])
        sheet.cell(row=5, column=6).value = max([j[1] for j in pixel_data])
        sheet.cell(row=6, column=6).value = 100 - pixel_std
        sheet.cell(row=7, column=6).value = (1 -(pixel_std/pixel_average))*100
        if len(pixel_number_filtered) != 0:
            sheet.cell(row=10, column=6).value = np.average(pixel_average_filtered)
            sheet.cell(row=11, column=6).value = np.std(pixel_average_filtered)
            sheet.cell(row=12, column=6).value = min(pixel_average_filtered)
            sheet.cell(row=13, column=6).value = max(pixel_average_filtered)
            sheet.cell(row=14, column=6).value = 100 - np.std(pixel_average_filtered)
            sheet.cell(row=15, column=6).value = (1 -(np.std(pixel_average_filtered)/np.average(pixel_average_filtered)))*100

        if len(not_counting) == 0:
            sheet.cell(row=16, column=5).value = 'Counting all pixels'
        else:
            sheet.cell(row=16, column=5).value = 'Not counting pixels:'
            # sheet.cell(row=17, column=5).value = ",".join([str(integer) for integer in not_counting])
            for i in range(0, len(not_counting)):
                sheet.cell(row=17+i, column=5).value = not_counting[i]


        sheet.cell(row=1, column=10).value = filename.replace('.JPG','').replace('.jpg','')
        img_excel = openpyxl.drawing.image.Image(new_img_file)
        img_excel.width = int(img_resized.shape[1])*0.2
        img_excel.height = int(img_resized.shape[0])*0.2
        sheet.add_image(img_excel, 'J2')
        sheet_overall.title = 'Overall'
        sheet_overall.cell(row=1, column=1).value = 'Vg (V)'
        sheet_overall.cell(row=1, column=2).value = 'Uniformity'
        sheet_overall.cell(row=1, column=3).value = 'Filtered Uniformity'
        sheet_overall.cell(row=k+2, column=1).value = int(Vg)
        sheet_overall.cell(row=k+2, column=2).value = (1 -(pixel_std/pixel_average))*100
        sheet_overall.cell(row=k+2, column=3).value = (1 -(np.std(pixel_average_filtered)/np.average(pixel_average_filtered)))*100
        k+=1


try:
    name_index = filename.upper().index('VD')
    sheet_overall.auto_filter.ref = "A1:C13"
    sheet_overall.auto_filter.add_sort_condition("A1:A13", True)
    wb.save(path+'/'+filename[:name_index].replace('_','').replace('.JPG','') + '.xlsx')

except ValueError:
    wb.save(path+'/Uniformity' + '.xlsx')




for file in os.listdir(path):
    if file.endswith('grid.jpg') and '-8V' not in file:
        os.remove(path+'/'+file)
